from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string

def xlsform(writer, sheet, cells = ["A", 1, "U", 15], fill = None, font = None, num_format = None, align = None, valign = None):
    ws = writer.sheets[sheet]
    col_begin = column_index_from_string(cells[0])
    row_begin = cells[1] 
    col_end = column_index_from_string(cells[2]) +1
    row_end = cells[3] +1 
    for c in range(col_begin, col_end):
        for z in range(row_begin, row_end):
            if fill != None:
                ws.cell(row=z, column=c).fill = fill
            if font != None:
                ws.cell(row=z, column=c).font = font
            if num_format != None:
                ws.cell(row=z, column=c).number_format = num_format
            if align != None:
                ws.cell(row=z, column=c).alignment =  Alignment(horizontal=align)
            if valign != None:
                ws.cell(row=z, column=c).alignment =  Alignment(vertical=valign)
                
def fill_palette(*args):
    pattern = []
    for c in args:
        p = PatternFill(start_color=c, end_color=c, fill_type='solid')
        pattern.append(p)
    return pattern

def get_font():
    font0 = Font(size=11, bold=True, color='000000') # Big Black
    font1 = Font(size=11, bold=True, color='FFFFFF') # Big White
    font2 = Font(size=10, bold=False, color='000000') # Small Black
    font3 = Font(size=10, bold=False, color='FFFFFF') # Small White
    return [font0, font1, font2, font3]
